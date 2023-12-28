from django.shortcuts import render, redirect, get_object_or_404, redirect
from django.http import HttpResponse, HttpResponseForbidden, JsonResponse, HttpResponseRedirect
from django.urls import reverse
from django.contrib.auth.decorators import login_required
from .forms import npp_form, fpp_form, edit_fpp_form, mpr_form, wtr_form, TicketForm,TicketReportForm, rng_form
import pyodbc
import json
from .models import NewPickingProblem, Ticket, ActionLog
from datetime import datetime
from openpyxl import Workbook
import logging
import random
from collections import Counter
from openpyxl.styles import Font, PatternFill
from PyPDF2 import PdfWriter, PdfReader
import io
import os
from django.conf import settings


@login_required
def menu_view(request):
    return render(request, 'menu.html')

@login_required
def pick_menu(request):
    return render(request, 'picking/pick_menu.html')

@login_required
def npp_view(request):
    if not request.user.is_superuser and not request.user.groups.filter(name__in=['Managers','Supervisors']).exists():
        message = "You don't have permission to access this feature."
        return render(request, 'forbidden.html', {'message': message})
    if request.method == 'POST':
        form = npp_form(request.user, request.POST)
        if form.is_valid():
            # Process the form data (e.g., save it to the database)
            new_pp = form.save(commit=False)
            new_pp.name = request.user.username
            new_pp.date = datetime.now().date()
            new_pp.time = datetime.now().time()
            job_number = request.POST.get('job_number')
            new_pp.job_number = job_number
            new_pp = form.save()
            reference_number = new_pp.reference_number  # Retrieve the generated reference number
            # Redirect to the success page with the reference number
            return redirect('success', reference_number=reference_number)
        else:
            form.errors
    else:
        form = npp_form(request.user)
    return render(request, 'picking/npp.html', {'form': form})

@login_required
def success_view(request, reference_number):
    return render(request, 'success.html', {'reference_number': reference_number})

def fetch_data(request):
        try:
            work_order = request.GET.get('work_order')
            connection = pyodbc.connect('DSN=AS400-PRODUCTION;'
                                        'UID=inqmxc;'
                                        'PWD=inqmxc123;'
                                        'Connection Timeout=30;'
                                        )
            with connection.cursor() as cursor:
                sql_query = """
                    SELECT DISTINCT TSCO, TSCODE, TSPN, TSREFN, TSTQTY, TSLOCF, USRPRF, TSDATE, TSTIME, TSUTMS, TSDFOR, FKITMSTR.IMDSC
                    FROM B20E386T.KBM400MFG.FKITSAVE FKITSAVE
                    LEFT JOIN B20E386T.KBM400MFG.FKITMSTR FKITMSTR
                    ON FKITSAVE.TSPN = FKITMSTR.IMPN
                    WHERE TSREFN = ?
                    AND TSCODE IN ('03', '09', '26')
                    AND TSTQTY > 0
                """
                cursor.execute(sql_query, (work_order,))

                # Fetch all rows from the query
                data = cursor.fetchall()

                # Create a list of dictionaries for results
                query_results = []
                for row in data:
                    result = {
                        'TSCO': row.TSCO,
                        'TSCODE': row.TSCODE,
                        'TSPN': row.TSPN,
                        'TSREFN': row.TSREFN,
                        'TSTQTY': row.TSTQTY,
                        'TSLOCF': row.TSLOCF,
                        'USRPRF': row.USRPRF,
                        'TSDATE': row.TSDATE,
                        'TSTIME': row.TSTIME,
                        'TSUTMS': row.TSUTMS,
                        'TSDFOR': row.TSDFOR,
                        'IMDSC': row.IMDSC
                    }
                    query_results.append(result)

                return JsonResponse({'results': query_results})

        except Exception as e:
            return JsonResponse({'error_message': str(e)}, status=500)

        finally:
        # Close the database connection in a finally block
            if connection:
                connection.close()

        return JsonResponse({'error_message': 'Invalid request'}, status=400)

@login_required
def fpp(request):
    # Retrieve all existing submissions ordered by reference number
    submissions = NewPickingProblem.objects.exclude(status='Finalized').order_by('reference_number')

    # Render the FPP page template with the submissions data
    return render(request, 'picking/fpp.html', {'submissions': submissions})

@login_required
def print_pp(request):
    # Retrieve all existing submissions ordered by reference number
    submissions = NewPickingProblem.objects.exclude(status='Pending').order_by('reference_number')

    # Render the FPP page template with the submissions data
    return render(request, 'picking/print_pp.html', {'submissions': submissions})

@login_required
def follow_up_picking_problems(request):
    # Retrieve all records from the NewPickingProblem model
    submissions = NewPickingProblem.objects.all()

    # Pass the data to the template
    return render(request, 'master_data.html', {'submissions': submissions})



@login_required
def final_fpp(request, reference_number):
    if not request.user.is_superuser and not request.user.groups.filter(name__in=['Managers','Supervisors']).exists():
        message = "You don't have permission to access this feature."
        return render(request, 'forbidden.html', {'message': message})
    try:
        submission = NewPickingProblem.objects.get(reference_number=reference_number)
    except NewPickingProblem.DoesNotExist:
        # Handle the case where the specified reference_number does not exist
        return HttpResponse("Invalid reference number")

    work_order = submission.work_order
    part_number = submission.part_number
    
    if request.method == 'POST':
        form = fpp_form(request.POST, instance=submission)
        if form.is_valid():
            # Update the submission with finalization data
            
            submission=form.save()
            if submission.finalized:
                submission.finalized_date = datetime.now().date()
                submission.finalized_time = datetime.now().time()
                submission.finalized_username = request.user.username
                submission.status = 'Finalized'
                submission.save()
            return redirect('fpp_success_page')  # Redirect to success page

    else:
        # Populate the initial_data dictionary with the retrieved values
        initial_data = {
            'employee_username': '',  # Add your logic to retrieve this field
            'location': '',  # Add your logic to retrieve this field
            'transaction_date': '',  # Add your logic to retrieve this field
            'transaction_time': '',  # Add your logic to retrieve this field
            'applies': '',  # Add your logic to retrieve this field
            'root_cause': '',  # Add your logic to retrieve this field
            'final_reason': '',  # Add your logic to retrieve this field
            'final_comments': '',  # Add your logic to retrieve this field
            'status': '',  # Add your logic to retrieve this field
            'work_order': work_order,  # Populate work_order with the retrieved value
            'part_number': part_number,
        }
        print(initial_data)
        form = fpp_form(request.POST if request.POST else None, initial=initial_data)

    return render(request, 'picking/final_fpp.html', {'submission': submission, 'form': form})

@login_required
def fpp_success_page(request):
    return render(request, 'picking/fpp_success_page.html')

@login_required
def fpp_fetch_data(request):
        try:
            work_order = request.GET.get('work_order')
            part_number = request.GET.get('part_number')
            connection = pyodbc.connect('DSN=AS400-PRODUCTION;'
                                        'UID=inqmxc;'
                                        'PWD=inqmxc123;'
                                        'Connection Timeout=30;'
                                        )
            with connection.cursor() as cursor:
                sql_query = """
                    SELECT TSCO, TSCODE, TSPN, TSREFN, TSTQTY, TSLOCF, USRPRF, TSDATE, TSTIME, TSUTMS
                    FROM B20E386T.KBM400MFG.FKITSAVE FKITSAVE
                    WHERE TSREFN = ? AND TSPN = ?
                    AND TSCODE IN ('03', '09', '26')
                    AND TSTQTY > 0
                """
                cursor.execute(sql_query, (work_order, part_number))

                # Fetch all rows from the query
                data = cursor.fetchall()

                # Create a list of dictionaries for results
                query_results = []
                for row in data:
                    result = {
                        'TSCO': row.TSCO,
                        'TSCODE': row.TSCODE,
                        'TSPN': row.TSPN,
                        'TSREFN': row.TSREFN,
                        'TSTQTY': row.TSTQTY,
                        'TSLOCF': row.TSLOCF,
                        'USRPRF': row.USRPRF,
                        'TSDATE': row.TSDATE,
                        'TSTIME': row.TSTIME,
                        'TSUTMS': row.TSUTMS,
                    }
                    query_results.append(result)

                return JsonResponse({'results': query_results})

        except Exception as e:
            return JsonResponse({'error_message': str(e)}, status=500)

@login_required
def edit_fpp(request,reference_number):
    if not request.user.is_superuser and not request.user.groups.filter(name__in=['Managers','Supervisors']).exists():
        message = "You don't have permission to access this feature."
        return render(request, 'forbidden.html', {'message': message})
    picking_problem = get_object_or_404(NewPickingProblem, reference_number=reference_number)
    if request.method == 'POST':
        form = edit_fpp_form(request.POST, instance=picking_problem)
        if form.is_valid():
            # Update the submission with finalization data
            
            form.save()
            return redirect('epp_success_page')  # Redirect to success page
        else:
            print("Form is not valid:", form.errors)
    else:
        form = edit_fpp_form(instance=picking_problem)

    context = {'form': form}
    return render(request, 'picking/edit_fpp.html', context)

@login_required
def epp_success_page(request):
    return render(request, 'picking/epp_success_page.html')

@login_required
def pp_reports_menu(request):
    return render(request, 'picking/pp_reports_menu.html')

@login_required
def mpr_menu(request):
    if request.method == 'POST':
        form = mpr_form(request.POST)
        if form.is_valid():
            # Get filtering criteria from the form
            finalized = form.cleaned_data['finalized']
            applied = form.cleaned_data['applied']
            app_dept = form.cleaned_data['app_dept']
            root_cause = form.cleaned_data['root_cause']
            line = form.cleaned_data['line']
            department = form.cleaned_data['department']
            start_date = form.cleaned_data['start_date']
            end_date = form.cleaned_data['end_date']
            shift = form.cleaned_data['shift']
            start_tdate = form.cleaned_data['start_tdate']
            end_tdate = form.cleaned_data['end_tdate']
            start_fdate = form.cleaned_data['start_fdate']
            end_fdate = form.cleaned_data['end_fdate']
            # Apply filters to the data
            problems = NewPickingProblem.objects.all()
            if finalized != '':
                problems = problems.filter(finalized=(finalized == 'Y'))
            
            if applied != '':
                if applied == 'Y':
                    problems = problems.filter(applied= 'Y')
                if applied == 'N':
                    problems = problems.filter(applied= 'N')
            
            if app_dept != '':
                problems = problems.filter(app_dept=app_dept)
            
            if root_cause != '':
                problems = problems.filter(root_cause=root_cause)
            
            if line != '':
                problems = problems.filter(line=line)
            
            if department != '':
                problems = problems.filter(department=department)
            
            if start_date != None:
                problems = problems.filter(date__gte=start_date)

            if end_date != None:
                problems = problems.filter(date__lte=end_date)
            
            if shift != '':
                problems = problems.filter(shift=shift)
            
            if start_tdate != None:
                problems = problems.filter(transaction_date__gte=start_tdate)
            
            if end_tdate != None:
                problems = problems.filter(transaction_date__lte=end_tdate)
            
            if start_fdate != None:
                problems = problems.filter(finalized_date__gte=start_fdate)
            
            if end_fdate != None:
                problems = problems.filter(finalized_date__lte=end_fdate)
            
            # Create an Excel workbook and populate it with data
        wb = Workbook()
        ws = wb.active
        ws.title = "Master Picking Report"
        header_fill = PatternFill(start_color='15418c', end_color='15418c', fill_type = 'solid')
        header_font = Font(bold=True, color='ffffff')
        # Create a header row in the Excel sheet
        header_row = ws.append(['Reference #', 'Status', 'W/O #', 'Job #','Shift Reported', 'Reporting Department', 
                'Reporting User', 'Comments', 'Date Reported', 'Time Reported', 'Line', 
                'Part', 'Defect', 'Quantity', 'Quantity Affected', 'UOM', 'Applied?', 
                'Root Cause', 'Responsible Department', 'Responsible User', 'Location', 
                'Transaction Date', 'Transaction Time', 'Final Comment', 'Date Finalized', 
                'Time Finalized'])
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        alt_fill1 = PatternFill(start_color='f2f2f2', end_color='f2f2f2', fill_type='solid')
        alt_fill2 = PatternFill(start_color='e5e5e5', end_color='e5e5e5', fill_type='solid')
        row_num = 2
        
        for problem in problems:
            ws.append([problem.reference_number, problem.status, problem.work_order, problem.job_number, problem.shift,
                    problem.department, problem.name, problem.comment, problem.date, problem.time,
                    problem.line, problem.part_number, problem.defect, problem.total_count,
                    problem.total_affected, problem.unit_type, problem.applied, problem.root_cause,
                    problem.app_dept, problem.employee_username, problem.location, problem.transaction_date,
                    problem.transaction_time, problem.final_comments, problem.finalized_date,
                    problem.finalized_time])
            for cell in ws[row_num]:
                cell.fill = alt_fill1 if row_num % 2 == 0 else alt_fill2
            row_num += 1

        # Create an HTTP response with the Excel file
        response = HttpResponse(content_type="application/ms-excel")
        response["Content-Disposition"] = "attachment; filename=MasterPickingReport.xlsx"

        # Save the workbook to the response
        wb.save(response)

        return response
    else:
        form = mpr_form()
    return render(request, 'reporting/mpr_menu.html', {'form': form})

@login_required
def ppdept_menu(request):
    if request.method == 'POST':
        form = mpr_form(request.POST)
        if form.is_valid():
            start_date = form.cleaned_data['start_date']
            end_date = form.cleaned_data['end_date']
            app_dept = form.cleaned_data['app_dept']
            applied = form.cleaned_data['applied']

            problems = NewPickingProblem.objects.all()

            if start_date:
                problems = problems.filter(date__gte=start_date)    
            if end_date:
                problems = problems.filter(date__lte=end_date)
            if app_dept:
                problems = problems.filter(app_dept=app_dept)
            if applied:
                if applied == 'Y':
                    problems = problems.filter(applied='Y')
                elif applied == 'N':
                    problems = problems.filter(applied='N')

            wb = Workbook()
            ws = wb.active
            ws.title = "Picking Problems by Department"

            ws.append(['Responsible Department', 'Date Reported', 'Reference #', 'Applied?', 'W/O', 'Location', 'Quantity', 'Quantity Affected',
                        'Transaction Date', 'Transaction Time', 'Root Cause'])
            for problem in problems:
                ws.append([problem.app_dept, problem.date, problem.reference_number, problem.applied, problem.work_order, problem.location, problem.total_count,
                            problem.total_affected, problem.transaction_date, problem.transaction_time, problem.root_cause])
                
            response = HttpResponse(content_type="application/ms-excel")
            response["Content-Disposition"] = "attachment; filename=PickingProblemsByDepartment.xlsx"

            wb.save(response)
            return response  # Return the Excel file response
    else:
        form = mpr_form()
    return render(request, 'reporting/ppdept_menu.html', {'form': form})

@login_required
def wtr_menu(request):
    if request.method == 'POST':
        form = wtr_form(request.POST)
        if form.is_valid():
            start_tdate = form.cleaned_data['start_tdate']
            end_tdate = form.cleaned_data['end_tdate']
            start_tdate = convert_date_format(start_tdate)
            end_tdate = convert_date_format(end_tdate)
            print(start_tdate)
            print(end_tdate)

        try:
            connection = pyodbc.connect('DSN=AS400-PRODUCTION;'
                                        'UID=inqmxc;'
                                        'PWD=inqmxc123;'
                                        'Connection Timeout=500;'
                                        )
            with connection.cursor() as cursor:
                sql_query = """
                    SELECT 
                    TSCO, TSCODE, TSPN, TSTQTY, TSLOCF, TSREFN, TSDFOR, 
                    TSLOCT, TSLOTN, USRPRF, TSDATE, TSTIME, TSRFLG, TSAINF
                    FROM B20E386T.KBM400MFG.FKITSAVE FKITSAVE
                    WHERE TSCODE IN ('03', '26') and FKITSAVE.TSDATE BETWEEN ? AND ?
                    AND TSCO = '710'
                """
                cursor.execute(sql_query,start_tdate,end_tdate)

                # Fetch all rows from the query
                data = cursor.fetchall()
                query_results = []

                if data:
                    response = HttpResponse(content_type="application/ms-excel")
                    response["Content-Disposition"] = "attachment; filename=WeeklyTransactionsReport.xlsx"

                    wb = Workbook()
                    ws = wb.active
                    ws.title = "Weekly Transactions Report"
                    header_fill = PatternFill(start_color='15418c', end_color='15418c', fill_type = 'solid')
                    header_font = Font(bold=True, color='ffffff')
                    # Create a header row in the Excel sheet
                    header = ['TSCO', 'TSCODE', 'TSPN', 'TSTQTY', 'TSLOCF', 'TSREFN', 'TSDFOR', 'TSLOCT',
                              'TSLOTN', 'USRPRF', 'TSDATE', 'TSTIME', 'TSRFLG', 'TSAINF']
                    header_row = ws.append(header)
                    for cell in ws[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                    alt_fill1 = PatternFill(start_color='f2f2f2', end_color='f2f2f2', fill_type='solid')
                    alt_fill2 = PatternFill(start_color='e5e5e5', end_color='e5e5e5', fill_type='solid')
                    row_num = 2
                    for row in data:
                        ws.append(list(row))
                        for cell in ws[row_num]:
                            cell.fill = alt_fill1 if row_num % 2 == 0 else alt_fill2
                        row_num += 1
                    wb.save(response)
                    return response
                else:
                    return HttpResponse("No data found for the specified date range.")
        except pyodbc.Error as e:
            return HttpResponse("An error occurred: " + str(e))
    else:
        form=mpr_form()
    return render(request, 'reporting/wtr_menu.html')


@login_required
def forbidden(request):
    if request.is_ajax():
        return HttpResponseForbidden('Access Denied')  # For AJAX requests
    return render(request, 'forbidden.html')  # Render a custom template for non-AJAX requests

@login_required
def eliminate_pp(request):
    if not request.user.is_superuser and not request.user.groups.filter(name__in=['Managers','Supervisors']).exists():
        message = "You don't have permission to access this feature."
        return render(request, 'forbidden.html', {'message': message})
    # Retrieve all records from the NewPickingProblem model
    submissions = NewPickingProblem.objects.filter(status = 'Pending').order_by('reference_number')
    if request.user.has_perm('ppLRPDN.can_delete_pp'):
        return render(request, 'picking/eliminate_pp.html', {'submissions': submissions})
    else:
        message = "You don't have permission to access this feature."
        return render(request, 'forbidden.html', {'message': message})

@login_required
def elim_pp(request,reference_number):
    if not request.user.is_superuser and not request.user.groups.filter(name__in=['Managers','Supervisors']).exists():
        message = "You don't have permission to access this feature."
        return render(request, 'forbidden.html', {'message': message})

    picking_problem = get_object_or_404(NewPickingProblem, reference_number=reference_number)

    if request.method == 'POST':
        if 'delete' in request.POST:
            # Delete the picking problem instance
            picking_problem.delete()
            ActionLog.objects.create(user=request.user, action_type='pp_deleted', reference_number=reference_number)
            return HttpResponseRedirect(reverse('eliminate_pp'))  # Redirect to the appropriate page

        # Handle the edit form
        form = edit_fpp_form(request.POST, instance=picking_problem)
        if form.is_valid():
            # Update the submission with finalization data
            form.save()
            return redirect('picking/epp_success_page')  # Redirect to the success page
        else:
            print("Form is not valid:", form.errors)
    else:
        form = edit_fpp_form(instance=picking_problem)

    context = {'form': form, 'picking_problem': picking_problem}
    return render(request, 'picking/elim_pp.html', context)

@login_required
def super_menu(request):
    if not request.user.is_superuser and not request.user.groups.filter(name__in=['Managers','Supervisors']).exists():
        message = "You don't have permission to access this feature."
        return render(request, 'forbidden.html', {'message': message})
    
    return render(request, 'super_menu.html')

logger= logging.getLogger(__name__)

@login_required
def create_ticket(request):
    if not request.user.is_superuser and not request.user.groups.filter(name__in=['Managers','Supervisors']).exists():
        message = "You don't have permission to access this feature."
        return render(request, 'forbidden.html', {'message': message})
    
    if request.method == 'POST':
        form = TicketForm(request.POST)

        if form.is_valid():
            new_ticket=form.save()
            # Redirect to a new URL if the ticket is saved successfully:
            ticket_number = new_ticket.ticket_number
            
            return redirect('ticket_success', ticket_number)  # Replace 'success_url' with the name of the URL to redirect to.
        else:
            return render(request, 'ticket/create_ticket.html', {'form': form, 'form_invalid': True})
    else:
        form = TicketForm()  # An unbound form
    return render(request, 'ticket/create_ticket.html', {'form': form})

@login_required
def ticket_success(request, ticket_number):
    return render(request, 'ticket/ticket_success.html', {'ticket_number': ticket_number})

@login_required
def ticket_menu(request):
    return render(request, 'ticket/ticket_menu.html')

@login_required
def ticket_list(request):
    submissions = Ticket.objects.order_by('ticket_number')
    return render(request, 'ticket/ticket_list.html', {'submissions': submissions})

@login_required
def edit_ticket(request, ticket_number):
    if not request.user.is_superuser and not request.user.groups.filter(name__in=['Managers','Supervisors']).exists():
        message = "You don't have permission to access this feature."
        return render(request, 'forbidden.html', {'message': message})
    ticket= get_object_or_404(Ticket, ticket_number=ticket_number)
    if request.method == 'POST':
        form = TicketForm(request.POST, instance=ticket)
        if form.is_valid():
            form.save()
            # Redirect to a new URL if the ticket is saved successfully:
            return redirect('edit_ticket_success', ticket_number)
    else:
        form = TicketForm(instance=ticket)
    context = {'form': form}
    return render(request, 'ticket/edit_ticket.html', context)

@login_required
def edit_ticket_success(request, ticket_number):
    return render(request, 'ticket/edit_ticket_success.html', {'ticket_number': ticket_number})

@login_required
def ticket_report(request):
    if request.method == 'POST':
        print("POST request received")
        form = TicketReportForm(request.POST)
        if form.is_valid():
            print("Form is valid")
            # Get filtering criteria from the form
            username = form.cleaned_data['username']
            supervisor = form.cleaned_data['supervisor']
            start_date = form.cleaned_data['start_date']
            end_date = form.cleaned_data['end_date']
            tickets=Ticket.objects.all()
            print(tickets)
            if username != '':
                tickets = tickets.filter(employee_name=username)
            if supervisor != '':
                tickets = tickets.filter(supervisor=supervisor)
            if start_date != None:
                tickets = tickets.filter(date__gte=start_date)
            if end_date != None:
                tickets = tickets.filter(date__lte=end_date)
            # Create an Excel workbook and populate it with data
            wb = Workbook()
            ws = wb.active
            ws.title = "Master Ticket Report"
            header_fill = PatternFill(start_color='15418c', end_color='15418c', fill_type = 'solid')
            header_font = Font(bold=True, color='ffffff')
            header = ws.append(['Ticket #', 'Username', 'Badge #', 'Date', 'Location', 'Item #', 
                        'Work Order', 'Item Description', 'Supplier', 'Point of Discovery', 
                        'Badge #', 'Non-Conformances', 'Lead Supervisor', 'Investigation Date', 
                        'Training', 'Training Date', 'Counseling', 'Counseling Date', 
                        'Disciplinary Action', 'Disciplinary Action Date', 'Notes', 
                        'Employee Signature', 'Employee Signature Date'])
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
            alt_fill1 = PatternFill(start_color='f2f2f2', end_color='f2f2f2', fill_type='solid')
            alt_fill2 = PatternFill(start_color='e5e5e5', end_color='e5e5e5', fill_type='solid')
            row_num = 2
            for ticket in tickets:
                ws.append([ticket.ticket_number, ticket.employee_name, ticket.badge_number, 
                        ticket.date, ticket.location, ticket.item_number, ticket.work_order,
                        ticket.item_description, ticket.supplier, ticket.point_of_discovery,
                        ticket.badge_number_pod, ', '.join(ticket.non_conformances) if isinstance(ticket.non_conformances, list)
                        else ticket.non_conformances, ticket.lead_supervisor,
                        ticket.investigation_date, ticket.training, ticket.training_date,
                        ticket.counseling, ticket.counseling_date, ticket.disciplinary_action,
                        ticket.disciplinary_action_date, ticket.notes, ticket.employee_signature,
                        ticket.employee_signature_date])
                for cell in ws[row_num]:
                    cell.fill = alt_fill1 if row_num % 2 == 0 else alt_fill2
                row_num += 1

            # Create an HTTP response with the Excel file
            response = HttpResponse(content_type="application/ms-excel")
            response["Content-Disposition"] = "attachment; filename=MasterTicketReport.xlsx"

            # Save the workbook to the response
            wb.save(response)

            return response
        else:
            print("Form is not valid:", form.errors)
    else:
        form = mpr_form()
    return render(request, 'reporting/ticket_report.html', {'form': form})

@login_required
def elim_ticket_list(request):
    submissions = Ticket.objects.order_by('ticket_number')
    return render(request, 'ticket/elim_ticket_list.html', {'submissions': submissions})

@login_required
def elim_ticket(request, ticket_number):
    if not request.user.is_superuser and not request.user.groups.filter(name__in=['Managers','Supervisors']).exists():
        message = "You don't have permission to access this feature."
        return render(request, 'forbidden.html', {'message': message})
    ticket = get_object_or_404(Ticket, ticket_number=ticket_number)
    if request.method == 'POST':
        if 'delete' in request.POST:
            # Delete the ticket instance
            ticket.delete()
            ActionLog.objects.create(user=request.user, action_type='ticket_deleted', reference_number=ticket_number)
            return HttpResponseRedirect(reverse('elim_ticket_list'))  # Redirect to the appropriate page

        # Handle the edit form
        form = TicketForm(request.POST, instance=ticket)
        if form.is_valid():
            # Update the submission with finalization data
            form.save()
            
            return redirect('edit_ticket_success', ticket_number)  # Redirect to the success page
        else:
            print("Form is not valid:", form.errors)
    else:
        form = TicketForm(instance=ticket)

    context = {'form': form, 'ticket': ticket}
    return render(request, 'ticket/elim_ticket.html', context)

def convert_date_format(date_obj):
    # Format the date into the format MMDDYY
    converted_date = date_obj.strftime('1'+'%y%m%d')
    return converted_date

@login_required
def rng_locations_menu(request):
    if request.method == 'POST':
        form = rng_form(request.POST)
        if form.is_valid():
            username = form.cleaned_data['username']
            start_date = form.cleaned_data['start_date']
            end_date = form.cleaned_data['end_date']
            count = form.cleaned_data['location_count']
            start_date = convert_date_format(start_date)
            end_date = convert_date_format(end_date)
            print(username)
            print(start_date)
            print(end_date)
            print(count)
            try:
                connection = pyodbc.connect('DSN=AS400-PRODUCTION;'
                                            'UID=inqmxc;'
                                            'PWD=inqmxc123;'
                                            'Connection Timeout=30;'
                                            )
                with connection.cursor() as cursor:
                    sql_query = """
                        SELECT 
                        TSCO, TSLOCF, USRPRF, TSDATE
                        FROM B20E386T.KBM400MFG.FKITSAVE FKITSAVE
                        WHERE TSCODE IN ('03', '26') and FKITSAVE.TSDATE BETWEEN ? AND ?
                        AND TSCO = '710' AND USRPRF = ?
                        """
                    cursor.execute(sql_query,start_date,end_date,username)

                    data = cursor.fetchall()
                    query_results = []
                    
                    if data:
                        selected_data = random.sample(data, min(len(data), count)) 

                        response= HttpResponse(content_type="application/ms-excel")
                        response["Content-Disposition"] = "attachment; filename=RandomLocations.xlsx"

                        wb = Workbook()
                        ws = wb.active
                        ws.title = "Random Locations"

                        header_fill = PatternFill(start_color='15418c', end_color='15418c', fill_type = 'solid')
                        header_font = Font(bold=True, color='ffffff')
                        header = ['TSCO', 'TSLOCF', 'USRPRF', 'TSDATE']
                        header_row=ws.append(header)
                        for cell in ws[1]:
                            cell.fill = header_fill
                            cell.font = header_font
                        alt_fill1 = PatternFill(start_color='f2f2f2', end_color='f2f2f2', fill_type='solid')
                        alt_fill2 = PatternFill(start_color='e5e5e5', end_color='e5e5e5', fill_type='solid')
                        row_num = 2
                        for row in selected_data:
                            ws.append(list(row))
                            for cell in ws[row_num]:
                                cell.fill = alt_fill1 if row_num % 2 == 0 else alt_fill2
                            row_num += 1
                        wb.save(response)
                        return response
                    else:
                        return HttpResponse("No data found for the specified date range.")               
            except pyodbc.Error as e:
                return HttpResponse("An error occurred: " + str(e))
        else:
            print("Form is not valid:", form.errors)
    else:
        
        form = rng_form()
    return render(request, 'ticket/rng_menu.html', {'form': form})

@login_required
def fer_menu(request):
    if request.method == 'POST':
        form = wtr_form(request.POST)
        if form.is_valid():
            start_tdate = form.cleaned_data['start_tdate']
            end_tdate = form.cleaned_data['end_tdate']
            start_tdate_converted = convert_date_format(start_tdate)
            end_tdate_converted = convert_date_format(end_tdate)
            print(start_tdate)
            print(end_tdate)

        try:
            connection = pyodbc.connect('DSN=AS400-PRODUCTION;'
                                        'UID=inqmxc;'
                                        'PWD=inqmxc123;'
                                        'Connection Timeout=30;'
                                        )
            with connection.cursor() as cursor:
                sql_query = """
                    SELECT 
                    TSCO, TSCODE, TSPN, TSTQTY, TSLOCF, TSREFN, TSDFOR, 
                    TSLOCT, TSLOTN, USRPRF, TSDATE, TSTIME, TSRFLG, TSAINF
                    FROM B20E386T.KBM400MFG.FKITSAVE FKITSAVE
                    WHERE TSCODE IN ('03', '26') and FKITSAVE.TSDATE BETWEEN ? AND ?
                    AND TSCO = '710'
                """
                cursor.execute(sql_query,start_tdate_converted,end_tdate_converted)

                # Fetch all rows from the query
                data = cursor.fetchall()
                usrprf_to_employee_username = {row.USRPRF: row.USRPRF for row in data}
                print(usrprf_to_employee_username)
                print(usrprf_to_employee_username.values())
                picking_problems_data= NewPickingProblem.objects.filter(
                    transaction_date__gte=start_tdate, 
                    transaction_date__lte=end_tdate,
                    employee_username__in=[usrprf.strip() for usrprf in usrprf_to_employee_username.values()]
                )
                print(picking_problems_data)
                picking_problems_count = Counter(p.employee_username for p in picking_problems_data)
                print(picking_problems_count)
                if data:
                    usrprf_counts = Counter(row.USRPRF for row in data)
                    response = HttpResponse(content_type="application/ms-excel")
                    response["Content-Disposition"] = "attachment; filename=FullEmployeeReport.xlsx"

                    wb = Workbook()
                    ws = wb.active
                    ws.title = "Full Employee Report"

                    
                    date_fill = PatternFill(start_color='728DBA', end_color='728DBA', fill_type='solid')
                    date_font = Font(bold=True, color='ffffff')
                    date_row = ws.append(['Start Date:', start_tdate, 'End Date:', end_tdate])
                    for cell in ws[1]:
                        cell.fill = date_fill
                        cell.font = date_font
                        

                    # Create a header row in the Excel sheet
                    header_fill = PatternFill(start_color='15418c', end_color='15418c', fill_type='solid')
                    header_font = Font(bold=True, color='ffffff')
                    header = ['USRPRF', 'PICKS', 'PICKING PROBLEMS', 'PP * 10k']
                    header_row = ws.append(header)
                    for cell in ws[2]:
                        cell.fill = header_fill
                        cell.font = header_font

                    alt_fill1 = PatternFill(start_color='f2f2f2', end_color='f2f2f2', fill_type='solid')
                    alt_fill2 = PatternFill(start_color='e5e5e5', end_color='e5e5e5', fill_type='solid')
                    row_num = 3
                    for usrprf, count in usrprf_counts.items():
                        pick_problems_count= picking_problems_count.get(usrprf.strip(), 0)
                        efficiency_metric = (pick_problems_count / count * 10000) if pick_problems_count != 0 else 0
                        ws.append([usrprf, count, pick_problems_count, efficiency_metric])

                        for cell in ws[row_num]:
                            cell.fill = alt_fill1 if row_num % 2 == 0 else alt_fill2
                        row_num += 1
                    wb.save(response)
                    return response
                else:
                    return HttpResponse("No data found for the specified date range.")
        except pyodbc.Error as e:
            return HttpResponse("An error occurred: " + str(e))
    else:
        form=mpr_form()
    return render(request, 'reporting/full_emp_report.html')

def fill_pdf(data):
    #Load the PDF
    template_path = os.path.join(settings.BASE_DIR, 'static\PickingProblems\pp_form.pdf')
    output = PdfWriter()
    input_pdf = PdfReader(open(template_path, "rb"))

    #Get the first page
    page = input_pdf.pages[0]
    fields = input_pdf.get_fields()

    
    #Update the form fields
    output.add_page(page)

    #Add the updated page to the writer
    output.update_page_form_field_values(output.pages[0],data)

    #Write the output PDF
    buffer = io.BytesIO()
    output.write(buffer)
    return buffer

def create_filled_pdf(request, reference_number):
    #Get problem instance
    problem = get_object_or_404(NewPickingProblem, reference_number=reference_number)
    date=datetime.strftime(problem.date, '%m/%d/%Y')
    finalized_date=datetime.strftime(problem.finalized_date, '%m/%d/%Y')
    #Define the data to fill in (keys must match the field names from the PDF form)
    data = {
        'reference_number' : problem.reference_number,
        'PART NUMBER' : problem.part_number,
        'ITEM DESCRIPTION' : problem.item_dsc,
        'PRODUCT #' : problem.job_number,
        'WORK ORDER #' : problem.work_order,
        'Picker #/Name' : problem.employee_username,
        'Issued By' : problem.name,
        'date' : date,
        'Received By' : problem.finalized_username,
        'finalized_date' : finalized_date,
        'Comments' : problem.final_comments,
    }

    print(data)
    pdf_buffer = fill_pdf(data)

    pdf_buffer.seek(0)

    response = HttpResponse(pdf_buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="PickingProblem_{reference_number}.pdf"'
    return response